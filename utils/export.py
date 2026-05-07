import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_excel(state, nvidia_results: dict) -> bytes:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    G  = "FF76B900"
    DK = "FF1A1A1A"
    WH = "FFFFFFFF"
    RD = "FFFF4444"
    GN = "FF44FF44"
    AM = "FFFFAA00"

    def hdr(ws, r, c, val, bg=DK, fc=WH):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="Arial", bold=True, color=fc, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return cell

    def val(ws, r, c, v, fc=DK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", color=fc, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return cell

    # Summary sheet
    ws_sum["A1"] = "BESS Simulation Summary"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="76B900")
    ws_sum["A2"] = "Tesla Megapack 2 XL | NVIDIA BESS Self-Qualification v0.4"
    ws_sum["A2"].font = Font(name="Arial", size=10, color="888888", italic=True)

    headers = ["Metric", "Value", "NVIDIA Limit", "Status"]
    for c, h in enumerate(headers, 1):
        hdr(ws_sum, 4, c, h)

    rows = [
        ("Peak IT Load (MW)", f"{state.peak_it_mw:.2f}", "—", "—"),
        ("Average IT Load (MW)", f"{state.avg_it_mw:.2f}", "—", "—"),
        ("Daily Energy (MWh)", f"{state.daily_mwh:.1f}", "—", "—"),
        ("Max Ramp Rate (% IT/s)", f"{state.max_ramp_pct_s:.1f}", "≤20%",
         "PASS" if state.max_ramp_pct_s <= 20 else "FAIL"),
        ("BESS MW Rated", f"{state.bess_mw_rated:.2f}", "—", "—"),
        ("BESS MWh Rated", f"{state.bess_mwh_rated:.2f}", "—", "—"),
        ("Tracking Error (%)", f"{state.tracking_err_pct:.2f}", "≤2%",
         "PASS" if state.tracking_err_pct <= 2 else "FAIL"),
        ("SOC Drift 24h (%)", f"{state.soc_drift_pct:+.1f}", "≤±5%",
         "PASS" if abs(state.soc_drift_pct) <= 5 else "FAIL"),
        ("Fast DR Response (s)", f"{state.dr_fast_resp_s:.1f}", "≤2s",
         "PASS" if state.dr_fast_resp_s <= 2 else "FAIL"),
        ("SCR", f"{state.scr:.1f}", "Test at SCR=2.0",
         "AMBER" if state.scr < 3 else "OK"),
        ("Generator MW Rated", f"{state.gen_mw_rated:.1f}", "—", "—"),
    ]
    for i, row in enumerate(rows):
        r = 5 + i
        for c, v in enumerate(row, 1):
            fg = DK
            if c == 4:
                fg = "006600" if v == "PASS" else ("CC0000" if v == "FAIL" else "885500")
            val(ws_sum, r, c, v, fg)

    # NVIDIA checks sheet
    ws_n = wb.create_sheet("NVIDIA Checks")
    for c, h in enumerate(["Test", "Name", "Method", "Status", "Detail"], 1):
        hdr(ws_n, 1, c, h)
    for i, (tid, res) in enumerate(nvidia_results.items()):
        r = 2 + i
        ws_n.cell(row=r, column=1, value=tid)
        ws_n.cell(row=r, column=2, value=res.get("name", ""))
        ws_n.cell(row=r, column=3, value=res.get("method", ""))
        ws_n.cell(row=r, column=4, value=res.get("status", ""))
        ws_n.cell(row=r, column=5, value=res.get("detail", ""))

    # Time series data sheet
    if state.time_s is not None:
        ws_ts = wb.create_sheet("Time Series")
        cols = {"Time_s": state.time_s}
        if state.p_it_mw   is not None: cols["P_IT_MW"]   = state.p_it_mw
        if state.p_bess_mw is not None: cols["P_BESS_MW"] = state.p_bess_mw
        if state.p_gen_mw  is not None: cols["P_GEN_MW"]  = state.p_gen_mw
        if state.p_grid_mw is not None: cols["P_GRID_MW"] = state.p_grid_mw
        if state.soc_pct   is not None: cols["SOC_PCT"]   = state.soc_pct
        if state.freq_hz   is not None: cols["FREQ_HZ"]   = state.freq_hz
        if state.v_pu      is not None: cols["V_PU"]      = state.v_pu

        # Downsample to 1-min to keep file manageable
        step = max(1, int(60 / state.dt_s))
        for c, (name, arr) in enumerate(cols.items(), 1):
            ws_ts.cell(row=1, column=c, value=name).font = Font(bold=True)
            for r, v in enumerate(arr[::step], 2):
                ws_ts.cell(row=r, column=c, value=round(float(v), 4))

    # Adjust column widths
    for ws in [ws_sum, ws_n]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
